import numpy as np
#import matplotlib.pyplot as plt
import pandas as pd
import Quandl as q #imports Quandl and assigns 'q' as its shortcut so we can just type q.get() etc
from openpyxl import Workbook
from openpyxl.styles import Style, Font, PatternFill, Color, Border, Side, Alignment


vix = q.get("YAHOO/INDEX_VIX")


monthlim = {1: 31 , 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}
xltomonth = {1: 'January',
	  	    2: 'February',
	   		3: 'March',
	   		4: 'April',
	   		5: 'May',
	   		6: 'June',
	   		7: 'July',
	   		8: 'August',
	   		9: 'September',
	   		10: 'October',
	   		11: 'November',
	   		12: 'December'}

xltomo = {1: 'JAN',
	  	    2: 'FEB',
	   		3: 'MAR',
	   		4: 'APR',
	   		5: 'MAY',
	   		6: 'JUN',
	   		7: 'JUL',
	   		8: 'AUG',
	   		9: 'SEP',
	   		10: 'OCT',
	   		11: 'NOV',
	   		12: 'DEC'}

# return a list with hex for spectrum from green to red
def colorlist():
	c = []
	for i in range(0, 104, 8): #12
		if i <= 8:
			c.append('0%XFF00' % i)
		else:
			c.append('%XFF00' % i)

	for j in range(223, 255, 8): # 4
		c.append('%XFF00' % j)
	for j in reversed(range(223, 255, 8)): #4
		c.append('FF%X00' % j)

	for k in reversed(range(0, 88, 8)): #
		if k > 8:
			c.append('FF%X00' % k)
		else:
			c.append('FF0%X00' % k)
	return c

#function to create a style from the passed cell color hex
def coloredcell(color):
	side = Side(border_style = 'medium', color = 'FF000000')
	xalignment = Alignment(horizontal = 'center')
	xborder = Border(left = side, right = side, top = side, bottom = side)
	xfont = Font(name='Calibri',
	                size=11,
	                bold=False,
	                italic=False,
	                vertAlign=None,
	                underline='none',
	                strike=False,
	                color='FF000000')
	return Style(font = xfont, fill = PatternFill(fill_type = 'solid', start_color = color, end_color = color), border = xborder, alignment = xalignment)

def comparisonmap(start, hold, buy, end):
	greenlist = np.linspace(start, hold, 14) # numpy linear space
	yellowlist = np.linspace(hold, buy, 8)	 # linspace(a,b,c) returns c evenly distributed values
	redlist = np.linspace(buy, end, 12)		 # from a to b, including endpoints a and b
	# concatenate the three lists and call unique() to get rid of duplicates
	compmap = np.unique(np.concatenate((greenlist, yellowlist, redlist)))
	return compmap

def setcell(sheet, row, column, value, style):
	sheet.cell(row = row, column = column).value = value
	sheet.cell(row = row, column = column).style = style

def printexcelmonthly(startyear, endyear, sheetname, start, hold, buy, end):

	xlbook = Workbook()
	xsheet = xlbook.active
	xsheet.title = sheetname
	colors = colorlist()
	comparisonlist = comparisonmap(start, hold, buy, end)
	length = len(comparisonlist)
	empty = coloredcell('FFFFFF')

	# make a style for every element of colors using coloredcell
	styles = [coloredcell(color) for color in colors] 

	# look up style with color corresponding to passed value
	def getstyle(value):
		for ceiling in comparisonlist:
			if value < ceiling: 
				z = np.where(comparisonlist==ceiling)
				return styles[z[0]]
			elif value >= comparisonlist[length - 1]: # we need to check if value is greater than upper bound too
				return styles[length - 1]

	# create the cells for January-December
	for month in range(12):
		setcell(xsheet, 1, month+2, xltomonth[month+1], empty)
		# xsheet.cell(row = 1, column = month+2).value = xltomonth[month+1]
		# xsheet.cell(row = 1, column = month+2).style = empty

	for year in range(startyear, endyear):
		xlrow = year - startyear + 2
		setcell(xsheet, xlrow, 1, xltomonth[month+1], "%s, %d" % (xltomonth[month+1], year), empty)
		# xsheet.cell(row = xlrow, column = 1).value = "%s, %d" % (xltomonth[month+1], year)
		# xsheet.cell(row = xlrow, column = 1).value = "%s, %d" % (xltomonth[month+1], year)
		for month in range(12):
			for day in range(monthlim[month + 1]):
				if year == startyear:
					day += 1
					curx = 999
				curdatestr = '%d-%d-%d' % (year, month + 1, day + 1)
				lastopen = vix.index.asof(curdatestr)
				curdate = pd.to_datetime(curdatestr)
				if curdate.is_month_start:
					curx = 999
				nextx = vix.loc[lastopen, 'Low']
				if nextx < curx:
					xdate = curdate
					curx = nextx
				if curdate.is_month_end: #we know we found the lowest now, so break
					break
		
			setcell(xsheet, xlrow, month+2, curx, getstyle(curx))
			# xsheet.cell(row = xlrow, column = month+2).value = curx #add value to cell
			# xsheet.cell(row = xlrow, column = month+2).style = getstyle(curx)
		
	xlbook.save('%s.xlsx' % sheetname)


def printexceldaily(startyear, endyear, sheetname, start, hold, buy, end):

	xlbook = Workbook()
	xsheet = xlbook.active
	xsheet.title = sheetname
	colors = colorlist()
	comparisonlist = comparisonmap(start, hold, buy, end)
	length = len(comparisonlist)
	empty = coloredcell('FFFFFF')

	# fill a list with a style for every element of colors using coloredcell
	styles = [coloredcell(color) for color in colors] 

	def getstyle(value):
		for ceiling in comparisonlist:
			if value < ceiling: 
				z = np.where(comparisonlist==ceiling)
				return styles[z[0]]
			elif value >= comparisonlist[length - 1]: # we need to check if value is greater than upper bound too
				return styles[length - 1]

	# create the cells for days
	for i in range(2, 32):
		setcell(xsheet, 1, i, i, empty)
		# xsheet.cell(row = 1, column = i).value = i
		# xsheet.cell(row = 1, column = i).style = empty
	
	xlrow = 2
	for year in range(startyear, endyear):
		for month in range(12):
			setcell(xsheet, xlrow, 1, "%s, %d" % (xltomo[month+1], year), empty)
			# xsheet.cell(row = xlrow, column = 1).value = "%s, %d" % (xltomo[month+1], year)
			for day in range(monthlim[month + 1]):
				curdatestr = '%d-%d-%d' % (year, month + 1, day + 1)
				try:
					curx = vix.loc[curdatestr, 'Low']
				except KeyError:
					setcell(xsheet, xlrow, day+2, 'Closed', empty)
					# xsheet.cell(row = xlrow, column = day+2).value = 'Closed'
					# xsheet.cell(row = xlrow, column = day+2).style = empty
				else:
					setcell(xsheet, xlrow, day+2, curx, getstyle(curx))
					# xsheet.cell(row = xlrow, column = day+2).value = curx #add value to cell
					# xsheet.cell(row = xlrow, column = day+2).style = getstyle(curx) #add style to cell
			xlrow += 1

	xlbook.save('%s.xlsx' % sheetname)

printexceldaily(1990, 2015, 'Daily_Lows', 10, 15, 18, 28)
